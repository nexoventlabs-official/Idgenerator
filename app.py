"""
Voter ID Card Generator v4.0 - Web App
========================================
- User  : /       Enter Epic Number + optional photo -> Generate & Download Card
- Admin : /admin  Import XLSX/CSV, view voters & generation stats

Database : MongoDB Atlas
Photos   : Cloudinary (user uploads)
Cards    : Cloudinary (generated_cards folder)
"""

import csv
import io
import json
import os
import random
import re
import secrets
import sys
import time
import threading
from datetime import datetime, timezone

from flask import (
    Flask, Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, session
)
from PIL import Image
from werkzeug.utils import secure_filename

import requests as http_requests
import cloudinary
import cloudinary.uploader
import cloudinary.api
from pymongo import MongoClient, ASCENDING, DESCENDING

try:
    import redis as _redis_lib
except ImportError:
    _redis_lib = None

import string

import config
from generate_cards import (
    setup_logging, generate_card, generate_serial_number,
    load_bold_font, get_text_width, load_member_photo
)
from security_fixes import (
    hash_pin, verify_pin, rate_limit, rate_limiter,
    validate_mobile, validate_epic, validate_pin, validate_otp,
    sanitize_search, validate_file_upload, login_tracker
)
from health_check import health_bp
from cloudinary_secure import generate_signed_url, generate_download_url
from face_detection import validate_photo_for_id_card, detect_face_in_image

# ══════════════════════════════════════════════════════════════════
#  PHASE 1: CELERY & CIRCUIT BREAKERS
# ══════════════════════════════════════════════════════════════════

# Import Celery tasks
from tasks import celery, generate_card_async

# Circuit breakers for external services
from pybreaker import CircuitBreaker

# Cloudinary circuit breaker
cloudinary_breaker = CircuitBreaker(
    fail_max=5,  # Open circuit after 5 failures
    reset_timeout=60,  # Keep circuit open for 60 seconds
    name='cloudinary'
)

# SMS API circuit breaker
sms_breaker = CircuitBreaker(
    fail_max=3,
    reset_timeout=120,
    name='sms_api'
)

# ── App Setup ────────────────────────────────────────────────────
app = Flask(__name__,
            template_folder=os.path.join(config.BASE_DIR, 'templates'),
            static_folder=os.path.join(config.BASE_DIR, 'static'))
app.secret_key = os.getenv('FLASK_SECRET', 'voter-id-gen-secret-2026')
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB max upload

# ══════════════════════════════════════════════════════════════════
#  PHASE 1: REDIS-BASED RATE LIMITER
# ══════════════════════════════════════════════════════════════════

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# Replace in-memory rate limiter with Redis-based limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    storage_uri=os.getenv('REDIS_URL', 'redis://localhost:6379/0'),
    default_limits=["200 per day", "50 per hour"],
    storage_options={"socket_connect_timeout": 30},
    strategy="fixed-window"
)

ALLOWED_IMG = {'png', 'jpg', 'jpeg', 'bmp'}
ALLOWED_DATA = {'xlsx', 'xls', 'csv'}

logger = setup_logging()

# ══════════════════════════════════════════════════════════════════
#  PHASE 2: REDIS SESSION STORE (For Horizontal Scaling)
# ══════════════════════════════════════════════════════════════════

from flask_session import Session

# Configure Redis-based session storage for multi-instance deployment
app.config['SESSION_TYPE'] = 'redis'
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'voter_session:'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours

# Connect to Redis for sessions
if os.getenv('REDIS_URL'):
    import redis
    app.config['SESSION_REDIS'] = redis.from_url(
        os.getenv('REDIS_URL'),
        decode_responses=False  # Keep binary for session data
    )
    Session(app)
    logger.info("Redis session store configured for horizontal scaling")
else:
    logger.warning("REDIS_URL not set - using default Flask sessions (not suitable for multi-instance)")

# ══════════════════════════════════════════════════════════════════
#  SECURITY: HTTPS ENFORCEMENT & CORS
# ══════════════════════════════════════════════════════════════════

# SECURITY FIX: Enforce HTTPS in production
from flask_talisman import Talisman
if os.getenv('FLASK_ENV') != 'development':
    Talisman(app, 
             force_https=True,
             strict_transport_security=True,
             strict_transport_security_max_age=31536000,
             content_security_policy={
                 'default-src': ["'self'", 'https://res.cloudinary.com', 'https://2factor.in'],
                 'img-src': ["'self'", 'https://res.cloudinary.com', 'data:'],
                 'style-src': ["'self'", "'unsafe-inline'", 'https://fonts.googleapis.com', 'https://cdn.jsdelivr.net'],
                 'script-src': ["'self'", "'unsafe-inline'", 'https://cdn.jsdelivr.net'],
                 'font-src': ["'self'", 'https://fonts.gstatic.com', 'https://cdn.jsdelivr.net'],
                 'connect-src': ["'self'", 'https://cdn.jsdelivr.net']
             })

# SECURITY FIX: Configure CORS for API endpoints
from flask_cors import CORS
CORS(app, resources={
    r"/api/*": {
        "origins": os.getenv('ALLOWED_ORIGINS', '*').split(','),
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# ══════════════════════════════════════════════════════════════════
#  SECURITY HEADERS & CDN CACHE CONTROL
# ══════════════════════════════════════════════════════════════════

@app.after_request
def set_security_headers(response):
    """SECURITY FIX: Add security headers to all responses."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # PHASE 2: Add cache control headers for CDN
    # Static assets get long cache
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    # API responses should not be cached
    elif request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    # HTML pages get short cache
    else:
        response.headers['Cache-Control'] = 'public, max-age=300'  # 5 minutes
    
    return response

for d in [config.MEMBER_PHOTOS_DIR, config.DATA_DIR, config.UPLOADS_DIR]:
    os.makedirs(d, exist_ok=True)

# ── IST Timezone Filter ──────────────────────────────────────
from datetime import timedelta
IST = timezone(timedelta(hours=5, minutes=30))

@app.template_filter('to_ist')
def to_ist(dt_str):
    """Convert a UTC ISO timestamp string to IST formatted string."""
    if not dt_str:
        return '-'
    try:
        s = str(dt_str).replace('Z', '+00:00')
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        ist_dt = dt.astimezone(IST)
        return ist_dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        return str(dt_str)[:19].replace('T', ' ') if dt_str else '-'

ALLOWED_IMG = {'png', 'jpg', 'jpeg', 'bmp'}
ALLOWED_DATA = {'xlsx', 'xls', 'csv'}

# Logger already initialized above

# ── MongoDB Setup (Main DB - voter data from XLSX imports, READ-ONLY after import) ──
# PHASE 2: Configure read preference for read replicas
from pymongo import ReadPreference

mongo_client = MongoClient(
    config.MONGO_URI,
    serverSelectionTimeoutMS=5000,
    maxPoolSize=200,         # SCALABILITY FIX: Increased from 50 to handle more concurrent connections
    minPoolSize=20,          # SCALABILITY FIX: Increased from 5 for better burst handling
    maxIdleTimeMS=30000,
    connectTimeoutMS=10000,
    socketTimeoutMS=300000,  # 5 min for large bulk imports
    # PHASE 2: Use secondary preferred for read-heavy operations
    readPreference='secondaryPreferred',
    # PHASE 2: Enable retryable reads for better reliability
    retryReads=True,
    retryWrites=True,
)
db = mongo_client[config.MONGO_DB_NAME]
voters_col = db[config.MONGO_VOTERS_COLLECTION]

# ── MongoDB Setup (Generated Voters DB - all card-generation activity) ──
gen_mongo_client = MongoClient(
    config.GEN_MONGO_URI,
    serverSelectionTimeoutMS=5000,
    maxPoolSize=200,         # SCALABILITY FIX: Increased from 50
    minPoolSize=20,          # SCALABILITY FIX: Increased from 5
    maxIdleTimeMS=30000,
    connectTimeoutMS=5000,
    socketTimeoutMS=10000,
    # PHASE 2: Use secondary preferred for read-heavy operations
    readPreference='secondaryPreferred',
    # PHASE 2: Enable retryable reads/writes
    retryReads=True,
    retryWrites=True,
)
gen_db = gen_mongo_client[config.GEN_MONGO_DB_NAME]
gen_voters_col = gen_db[config.GEN_MONGO_COLLECTION]
stats_col = gen_db[config.MONGO_STATS_COLLECTION]
verified_mobiles_col = gen_db['verified_mobiles']
otp_col = gen_db['otp_sessions']
volunteer_requests_col = gen_db['volunteer_requests']
booth_agent_requests_col = gen_db['booth_agent_requests']

# ── Background import state ──
import_status = {
    'running': False,
    'phase': '',        # 'parsing', 'inserting', 'done', 'error'
    'processed': 0,
    'inserted': 0,
    'total': 0,         # 0 until known
    'message': '',
    'error': '',
}
_import_lock = threading.Lock()

# Ensure indexes (graceful - don't crash if Atlas is unreachable)
try:
    # 1st MongoDB - voter data only
    voters_col.create_index('epic_no', unique=True)
    # Compound indexes for filter & search at scale
    voters_col.create_index([('assembly', 1), ('district', 1)])
    voters_col.create_index([('name', 1)])
    # Text index for fast $text search across multiple fields
    try:
        voters_col.create_index(
            [('name', 'text'), ('epic_no', 'text'), ('assembly', 'text'), ('district', 'text')],
            name='voters_text_search',
            default_language='none',
        )
    except Exception:
        pass  # Text index may already exist with different spec
    logger.info("MongoDB (voters) connected & indexes ensured.")
except Exception as e:
    logger.warning(f"MongoDB (voters) index creation skipped: {e}")

try:
    # 2nd MongoDB - all generation activity
    stats_col.create_index('epic_no', unique=True)
    stats_col.create_index([('count', 1)])
    stats_col.create_index([('auth_mobile', 1)])  # SCALABILITY FIX: Index for PIN verification
    gen_voters_col.create_index('ptc_code', unique=True)
    gen_voters_col.create_index('epic_no')
    gen_voters_col.create_index('mobile')
    gen_voters_col.create_index('referral_id', unique=True, sparse=True)
    gen_voters_col.create_index('referred_by_ptc')
    # CONCURRENCY FIX: Unique compound index to prevent race conditions
    gen_voters_col.create_index([('epic_no', 1), ('mobile', 1)], unique=True, name='epic_mobile_unique')
    # Compound indexes for filter, search & sort at scale
    gen_voters_col.create_index([('assembly', 1), ('district', 1)])
    gen_voters_col.create_index([('generated_at', -1)])
    gen_voters_col.create_index([('name', 1)])
    # Text index for fast $text search across multiple fields
    try:
        gen_voters_col.create_index(
            [('name', 'text'), ('epic_no', 'text'), ('ptc_code', 'text'), ('mobile', 'text'), ('assembly', 'text'), ('district', 'text')],
            name='gen_voters_text_search',
            default_language='none',
        )
    except Exception:
        pass  # Text index may already exist with different spec
    otp_col.create_index('mobile', unique=True)
    verified_mobiles_col.create_index('mobile', unique=True)
    volunteer_requests_col.create_index('ptc_code', unique=True)
    volunteer_requests_col.create_index('mobile')
    volunteer_requests_col.create_index('status')
    # Compound indexes for status + sort
    volunteer_requests_col.create_index([('status', 1), ('requested_at', -1)])
    volunteer_requests_col.create_index([('name', 1)])
    booth_agent_requests_col.create_index('ptc_code', unique=True)
    booth_agent_requests_col.create_index('mobile')
    booth_agent_requests_col.create_index('status')
    # Compound indexes for status + sort
    booth_agent_requests_col.create_index([('status', 1), ('requested_at', -1)])
    booth_agent_requests_col.create_index([('name', 1)])
    logger.info("MongoDB (generated) connected & indexes ensured.")
except Exception as e:
    logger.warning(f"MongoDB (generated) index creation skipped: {e}")

# ── Cloudinary Setup ─────────────────────────────────────────────
cloudinary.config(
    cloud_name=config.CLOUDINARY_CLOUD_NAME,
    api_key=config.CLOUDINARY_API_KEY,
    api_secret=config.CLOUDINARY_API_SECRET,
    secure=True,
)


# ══════════════════════════════════════════════════════════════════
#  REDIS CACHE SETUP  (P2 - optional, graceful degradation)
# ══════════════════════════════════════════════════════════════════

_redis_client = None
REDIS_DASHBOARD_KEY = 'voter_app:dashboard_stats'
REDIS_DASHBOARD_TTL = 60  # seconds
REDIS_EXTERNAL_STATS_KEY = 'voter_app:external_stats'
REDIS_EXTERNAL_STATS_TTL = 300  # 5 min for Cloudinary/SMS/dbstats (slow HTTP calls)
REDIS_DROPDOWN_VOTERS_KEY = 'voter_app:dropdown:voters'
REDIS_DROPDOWN_GEN_KEY = 'voter_app:dropdown:gen_voters'
REDIS_DROPDOWN_TTL = 300  # 5 min - assembly/district lists change very rarely

try:
    _redis_url = os.getenv('REDIS_URL', 'redis://localhost:6379/0')
    if _redis_lib:
        _redis_client = _redis_lib.from_url(_redis_url, socket_connect_timeout=2, decode_responses=True)
        _redis_client.ping()
        logger.info(f"Redis connected: {_redis_url}")
except Exception as _re:
    _redis_client = None
    logger.info(f"Redis not available (dashboard cache disabled): {_re}")


def _cache_get(key: str) -> dict | None:
    """Get a JSON value from Redis cache. Returns None on miss or if Redis unavailable."""
    if not _redis_client:
        return None
    try:
        raw = _redis_client.get(key)
        return json.loads(raw) if raw else None
    except Exception:
        return None


def _cache_set(key: str, value: dict, ttl: int = 60):
    """Set a JSON value in Redis cache with TTL. No-op if Redis unavailable."""
    if not _redis_client:
        return
    try:
        _redis_client.setex(key, ttl, json.dumps(value))
    except Exception:
        pass


# ══════════════════════════════════════════════════════════════════
#  MONGODB SEARCH / FILTER HELPERS  (for 5.6 Cr scale)
# ══════════════════════════════════════════════════════════════════

def _build_search_filter(search: str, fields: list[str]) -> dict:
    """Build a MongoDB $or filter using case-insensitive regex across fields."""
    if not search:
        return {}
    escaped = re.escape(search)
    regex = {'$regex': escaped, '$options': 'i'}
    return {'$or': [{f: regex} for f in fields]}


def _merge_conditions(conditions: list[dict]) -> dict:
    """Merge a list of query conditions into a single MongoDB query."""
    if not conditions:
        return {}
    if len(conditions) == 1:
        return conditions[0]
    return {'$and': conditions}


def _build_cursor_filter(cursor_id: str, direction: str = 'next', descending: bool = True) -> dict:
    """Build after/before filter for cursor-based pagination using _id.

    descending=True  (sort _id:-1): 'next' = $lt, 'prev' = $gt
    descending=False (sort _id:1):  'next' = $gt, 'prev' = $lt
    """
    if not cursor_id:
        return {}
    from bson import ObjectId
    try:
        oid = ObjectId(cursor_id)
    except Exception:
        return {}
    if descending:
        return {'_id': {'$lt': oid}} if direction == 'next' else {'_id': {'$gt': oid}}
    else:
        return {'_id': {'$gt': oid}} if direction == 'next' else {'_id': {'$lt': oid}}



# ══════════════════════════════════════════════════════════════════
#  FILE PARSING HELPERS (XLSX / CSV -> list of dicts)
# ══════════════════════════════════════════════════════════════════

def _match_column(header: str, candidates: list[str]) -> bool:
    h = header.strip().lower()
    for c in candidates:
        if c.lower() == h or c.lower() in h or h in c.lower():
            return True
    return False


def _safe_str(val):
    """Convert cell value to clean string, return '' for None."""
    if val is None:
        return ''
    return str(val).strip()


def _iter_xlsx(xlsx_path: str):
    """Generator: yield one voter dict at a time from an XLSX file (low memory)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else '')

    col_map = {}
    for field, candidates in config.XLSX_COLUMNS.items():
        for idx, h in enumerate(headers):
            if _match_column(h, candidates):
                col_map[field] = idx
                break

    mapped_indices = set(col_map.values())
    seen_epics = set()

    for row in ws.iter_rows(min_row=2):
        cells = [cell.value for cell in row]
        epic = _safe_str(cells[col_map['epic_no']] if 'epic_no' in col_map and col_map['epic_no'] < len(cells) else '')

        if not epic:
            continue

        epic_upper = epic.strip().upper()
        if epic_upper in seen_epics:
            continue
        seen_epics.add(epic_upper)

        voter = {'epic_no': epic}
        for field, idx in col_map.items():
            if field == 'epic_no':
                continue
            voter[field] = _safe_str(cells[idx] if idx < len(cells) else '')

        # Split combined lat_long into separate fields
        if 'lat_long' in voter and voter['lat_long']:
            parts = voter['lat_long'].split(',')
            if len(parts) == 2:
                voter['latitude'] = parts[0].strip()
                voter['longitude'] = parts[1].strip()
            del voter['lat_long']

        for idx, h in enumerate(headers):
            if idx not in mapped_indices and h:
                key = h.replace(' ', '_').lower()
                val = _safe_str(cells[idx] if idx < len(cells) else '')
                if val:
                    voter[key] = val

        yield voter

    wb.close()


def _iter_csv_bytes(raw: bytes):
    """Generator: yield one voter dict at a time from CSV bytes (low memory)."""
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode('latin-1')

    reader = csv.reader(io.StringIO(text))
    try:
        headers = [h.strip() for h in next(reader)]
    except StopIteration:
        return

    col_map = {}
    for field, candidates in config.XLSX_COLUMNS.items():
        for idx, h in enumerate(headers):
            if _match_column(h, candidates):
                col_map[field] = idx
                break

    mapped_indices = set(col_map.values())
    seen_epics = set()

    for cells in reader:
        epic = _safe_str(cells[col_map['epic_no']] if 'epic_no' in col_map and col_map['epic_no'] < len(cells) else '')
        if not epic:
            continue

        epic_upper = epic.strip().upper()
        if epic_upper in seen_epics:
            continue
        seen_epics.add(epic_upper)

        voter = {'epic_no': epic}
        for field, idx in col_map.items():
            if field == 'epic_no':
                continue
            voter[field] = _safe_str(cells[idx] if idx < len(cells) else '')

        # Split combined lat_long into separate fields
        if 'lat_long' in voter and voter['lat_long']:
            parts = voter['lat_long'].split(',')
            if len(parts) == 2:
                voter['latitude'] = parts[0].strip()
                voter['longitude'] = parts[1].strip()
            del voter['lat_long']

        for idx, h in enumerate(headers):
            if idx not in mapped_indices and h:
                key = h.replace(' ', '_').lower()
                val = _safe_str(cells[idx] if idx < len(cells) else '')
                if val:
                    voter[key] = val

        yield voter


# ══════════════════════════════════════════════════════════════════
#  MONGODB DATABASE HELPERS
# ══════════════════════════════════════════════════════════════════

def load_voters_from_db() -> list[dict]:
    """Return all voters from MongoDB."""
    return list(voters_col.find({}, {'_id': 0}))


def find_voter_by_epic(epic_no: str) -> dict | None:
    epic_no = epic_no.strip().upper()
    doc = voters_col.find_one({'epic_no': epic_no}, {'_id': 0})
    return doc


def _stream_upsert(voter_iter, status: dict) -> int:
    """Stream-upsert voters from an iterator into MongoDB in batches."""
    from pymongo import UpdateOne
    total = 0
    BATCH = 500
    batch = []
    for v in voter_iter:
        batch.append(v)
        status['processed'] += 1
        if len(batch) >= BATCH:
            ops = [UpdateOne({'epic_no': d['epic_no']}, {'$set': {k: val for k, val in d.items() if k != '_id'}}, upsert=True) for d in batch]
            result = voters_col.bulk_write(ops, ordered=False)
            total += result.upserted_count + result.modified_count
            status['inserted'] = total
            batch = []
    if batch:
        ops = [UpdateOne({'epic_no': d['epic_no']}, {'$set': {k: val for k, val in d.items() if k != '_id'}}, upsert=True) for d in batch]
        result = voters_col.bulk_write(ops, ordered=False)
        total += result.upserted_count + result.modified_count
        status['inserted'] = total
    return total


def _stream_replace(voter_iter, status: dict) -> int:
    """Drop existing voters and stream-insert from iterator in batches."""
    voters_col.delete_many({})
    total = 0
    BATCH = 500
    batch = []
    for v in voter_iter:
        batch.append(v)
        status['processed'] += 1
        if len(batch) >= BATCH:
            voters_col.insert_many(batch, ordered=False)
            total += len(batch)
            status['inserted'] = total
            batch = []
    if batch:
        voters_col.insert_many(batch, ordered=False)
        total += len(batch)
        status['inserted'] = total
    return total


# ══════════════════════════════════════════════════════════════════
#  GENERATION STATS (MongoDB)
# ══════════════════════════════════════════════════════════════════

def generate_ptc_code() -> str:
    """Generate a unique PTC-XXXXXXX code (collision-free under concurrent load)."""
    import uuid
    # Use UUID4 hex for guaranteed uniqueness - no DB check loop needed
    uid = uuid.uuid4().hex[:7].upper()
    return f'PTC-{uid}'


def save_generated_voter(voter: dict, mobile: str, photo_url: str, card_url: str, ptc_code: str,
                        referred_by_ptc: str = '', referred_by_referral_id: str = '',
                        secret_pin: str = ''):
    """Save a generated voter record to the new Generated Voters DB with atomic upsert."""
    import pymongo.errors
    
    doc = {
        'ptc_code': ptc_code,
        'epic_no': voter.get('epic_no', ''),
        'name': voter.get('name', ''),
        'assembly': voter.get('assembly', ''),
        'district': voter.get('district', ''),
        'mobile': mobile,
        'photo_url': photo_url,
        'card_url': card_url,
        'generated_at': datetime.now(timezone.utc).isoformat(),
    }
    if secret_pin:
        # SECURITY: Hash PIN before storing
        doc['secret_pin'] = hash_pin(secret_pin)
    if referred_by_ptc:
        doc['referred_by_ptc'] = referred_by_ptc
    if referred_by_referral_id:
        doc['referred_by_referral_id'] = referred_by_referral_id
    
    # CONCURRENCY FIX: Atomic upsert with duplicate key handling
    try:
        gen_voters_col.update_one(
            {'epic_no': voter.get('epic_no', ''), 'mobile': mobile},
            {'$set': doc, '$setOnInsert': {'created_at': datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
    except pymongo.errors.DuplicateKeyError:
        # Race condition: record already exists, just update it
        gen_voters_col.update_one(
            {'epic_no': voter.get('epic_no', ''), 'mobile': mobile},
            {'$set': doc}
        )
    
    # CONCURRENCY FIX: Atomic referral count increment
    if referred_by_ptc:
        gen_voters_col.update_one(
            {'ptc_code': referred_by_ptc},
            {'$inc': {'referred_members_count': 1}}
        )


def get_or_create_referral(ptc_code: str) -> dict | None:
    """Return { referral_id, referral_link } - idempotent."""
    voter = gen_voters_col.find_one({'ptc_code': ptc_code})
    if not voter:
        return None
    if voter.get('referral_id'):
        return {
            'referral_id': voter['referral_id'],
            'referral_link': f"{config.BASE_URL}/refer/{ptc_code}/{voter['referral_id']}"
        }
    import uuid
    rid = 'REF-' + uuid.uuid4().hex[:8].upper()
    link = f"{config.BASE_URL}/refer/{ptc_code}/{rid}"
    gen_voters_col.update_one(
        {'ptc_code': ptc_code},
        {'$set': {
            'referral_id': rid,
            'referral_link': link,
        },
         '$setOnInsert': {'referred_members_count': 0}}
    )
    # Ensure referred_members_count exists
    gen_voters_col.update_one(
        {'ptc_code': ptc_code, 'referred_members_count': {'$exists': False}},
        {'$set': {'referred_members_count': 0}}
    )
    return {'referral_id': rid, 'referral_link': link}


def generate_secret_pin() -> str:
    """Generate a cryptographically secure 4-digit PIN."""
    return f"{secrets.randbelow(10000):04d}"


def increment_generation_count(epic_no: str, photo_url: str = '', card_url: str = '',
                               auth_mobile: str = '', secret_pin: str = ''):
    """Increment generation count; optionally update photo_url, card_url, auth_mobile, secret_pin."""
    update = {
        '$inc': {'count': 1},
        '$set': {'last_generated': datetime.now(timezone.utc).isoformat()},
        '$setOnInsert': {'epic_no': epic_no},
    }
    if photo_url:
        update['$set']['photo_url'] = photo_url
    if card_url:
        update['$set']['card_url'] = card_url
    if auth_mobile:
        update['$set']['auth_mobile'] = auth_mobile
    if secret_pin:
        # SECURITY: Hash PIN before storing
        update['$set']['secret_pin'] = hash_pin(secret_pin)
    stats_col.update_one({'epic_no': epic_no}, update, upsert=True)


def get_voter_gen_count(epic_no: str) -> int:
    doc = stats_col.find_one({'epic_no': epic_no}, {'count': 1})
    return doc.get('count', 0) if doc else 0


def get_voter_photo_url(epic_no: str) -> str:
    """Get Cloudinary photo URL for a voter (from stats collection)."""
    doc = stats_col.find_one({'epic_no': epic_no}, {'photo_url': 1})
    return doc.get('photo_url', '') if doc else ''


def get_voter_card_url(epic_no: str) -> str:
    """Get Cloudinary card URL for a voter."""
    doc = stats_col.find_one({'epic_no': epic_no}, {'card_url': 1})
    return doc.get('card_url', '') if doc else ''


def get_all_stats() -> dict:
    """Return {epic_no: {count, last_generated, photo_url, card_url, auth_mobile}} dict."""
    result = {}
    for doc in stats_col.find({}, {'_id': 0}):
        result[doc['epic_no']] = {
            'count': doc.get('count', 0),
            'last_generated': doc.get('last_generated', ''),
            'photo_url': doc.get('photo_url', ''),
            'card_url': doc.get('card_url', ''),
            'auth_mobile': doc.get('auth_mobile', ''),
        }
    return result


# ══════════════════════════════════════════════════════════════════
#  CLOUDINARY UPLOAD
# ══════════════════════════════════════════════════════════════════

def upload_photo_to_cloudinary(image: Image.Image, epic_no: str) -> str:
    """Upload a PIL image to Cloudinary, return the secure URL."""
    buf = io.BytesIO()
    image.save(buf, format='JPEG', quality=90)
    buf.seek(0)

    result = cloudinary.uploader.upload(
        buf,
        folder=config.CLOUDINARY_PHOTO_FOLDER,
        public_id=epic_no,
        overwrite=True,
        resource_type='image',
    )
    return result.get('secure_url', '')


def upload_card_to_cloudinary(card_image: Image.Image, epic_no: str) -> str:
    """Upload generated card image to Cloudinary (overwrite mode - no delete needed)."""
    safe_id = epic_no.replace('/', '_').replace('\\', '_')

    buf = io.BytesIO()
    card_image.save(buf, format='JPEG', quality=config.JPEG_QUALITY)
    buf.seek(0)

    result = cloudinary.uploader.upload(
        buf,
        folder=config.CLOUDINARY_CARDS_FOLDER,
        public_id=safe_id,
        overwrite=True,
        invalidate=True,
        resource_type='image',
    )
    url = result.get('secure_url', '')
    logger.info(f"Uploaded card to Cloudinary: {url}")
    return url


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════

def allowed_file(filename, exts):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in exts


def _get_external_stats() -> dict:
    """Fetch slow external stats (Cloudinary, SMS, dbstats) - cached 5 min."""
    cached = _cache_get(REDIS_EXTERNAL_STATS_KEY)
    if cached and 'db1_size_mb' in cached:
        return cached

    result = {'mongodb_size_mb': 0, 'db1_size_mb': 0, 'db2_size_mb': 0,
              'db2_storage_mb': 0, 'db2_collections': 0, 'db2_objects': 0,
              'cloudinary_credits': 'N/A', 'sms_balance': 'N/A'}
    try:
        db_stats_1 = db.command('dbstats')
        db_stats_2 = gen_db.command('dbstats')
        d1 = db_stats_1.get('dataSize', 0)
        d2 = db_stats_2.get('dataSize', 0)
        result['mongodb_size_mb'] = round((d1 + d2) / (1024 * 1024), 2)
        result['db1_size_mb'] = round(d1 / (1024 * 1024), 2)
        result['db2_size_mb'] = round(d2 / (1024 * 1024), 2)
        result['db2_storage_mb'] = round(db_stats_2.get('storageSize', 0) / (1024 * 1024), 2)
        result['db2_collections'] = db_stats_2.get('collections', 0)
        result['db2_objects'] = db_stats_2.get('objects', 0)
    except Exception:
        pass

    try:
        cli_usage = cloudinary.api.usage()
        used = cli_usage.get('credits', {}).get('usage', 0)
        result['cloudinary_credits'] = f"{round(used, 2)}"
    except Exception:
        pass

    sms_api_key = os.getenv('SMS_API_KEY', '')
    if sms_api_key:
        try:
            resp = http_requests.get(f"https://2factor.in/API/V1/{sms_api_key}/BAL/SMS", timeout=3)
            if resp.status_code == 200:
                result['sms_balance'] = resp.json().get('Details', 'N/A')
        except Exception:
            pass

    _cache_set(REDIS_EXTERNAL_STATS_KEY, result, REDIS_EXTERNAL_STATS_TTL)
    return result


def _get_cached_dropdowns(collection, cache_key: str) -> tuple[list, list]:
    """Return (assemblies, districts) from cache or distinct() scan."""
    cached = _cache_get(cache_key)
    if cached:
        return cached.get('assemblies', []), cached.get('districts', [])
    assemblies = sorted(a for a in collection.distinct('assembly') if a)
    districts = sorted(d for d in collection.distinct('district') if d)
    _cache_set(cache_key, {'assemblies': assemblies, 'districts': districts}, REDIS_DROPDOWN_TTL)
    return assemblies, districts


def get_dashboard_stats():
    # P2: Try Redis cache first - instant dashboard load
    cached = _cache_get(REDIS_DASHBOARD_KEY)
    if cached:
        return cached

    try:
        # Use estimated_document_count - O(1) via collection metadata (fast at 5.6 Cr)
        total_voters = voters_col.estimated_document_count()

        # P1: Single aggregation for stats_col (1 query instead of loading ALL docs)
        stats_agg = list(stats_col.aggregate([
            {'$group': {
                '_id': None,
                'total_generated': {'$sum': {'$cond': [{'$gt': ['$count', 0]}, 1, 0]}},
                'total_generations': {'$sum': '$count'},
                'cards_on_cloud': {'$sum': {'$cond': [{'$gt': ['$card_url', '']}, 1, 0]}},
            }}
        ]))
        if stats_agg:
            total_generated = stats_agg[0].get('total_generated', 0)
            total_generations = stats_agg[0].get('total_generations', 0)
            cards_on_cloud = stats_agg[0].get('cards_on_cloud', 0)
        else:
            total_generated = 0
            total_generations = 0
            cards_on_cloud = 0

        db_connected = True

        # Generated voters count - O(1)
        try:
            generated_voters_count = gen_voters_col.estimated_document_count()
        except Exception:
            generated_voters_count = 0

        # P1: Single aggregation for gen_voters referral total
        try:
            ref_agg = list(gen_voters_col.aggregate([
                {'$group': {
                    '_id': None,
                    'total_referrals': {'$sum': {'$ifNull': ['$referred_members_count', 0]}},
                }}
            ]))
            total_referrals = ref_agg[0]['total_referrals'] if ref_agg else 0
        except Exception:
            total_referrals = 0

        # P1: Single aggregation for volunteer status counts
        try:
            vol_agg = list(volunteer_requests_col.aggregate([
                {'$group': {'_id': '$status', 'count': {'$sum': 1}}}
            ]))
            vol_counts = {doc['_id']: doc['count'] for doc in vol_agg}
            pending_volunteers = vol_counts.get('pending', 0)
            confirmed_volunteers = vol_counts.get('confirmed', 0)
        except Exception:
            pending_volunteers = 0
            confirmed_volunteers = 0

        # P1: Single aggregation for booth agent status counts
        try:
            ba_agg = list(booth_agent_requests_col.aggregate([
                {'$group': {'_id': '$status', 'count': {'$sum': 1}}}
            ]))
            ba_counts = {doc['_id']: doc['count'] for doc in ba_agg}
            pending_booth_agents = ba_counts.get('pending', 0)
            confirmed_booth_agents = ba_counts.get('confirmed', 0)
        except Exception:
            pending_booth_agents = 0
            confirmed_booth_agents = 0

    except Exception:
        total_voters = 0
        total_generated = 0
        total_generations = 0
        cards_on_cloud = 0
        generated_voters_count = 0
        db_connected = False
        total_referrals = 0
        pending_volunteers = 0
        confirmed_volunteers = 0
        pending_booth_agents = 0
        confirmed_booth_agents = 0

    # External stats loaded via AJAX separately (/api/external-stats)
    # Don't block dashboard render on slow external calls

    result = {
        'total_voters': total_voters,
        'total_generated': total_generated,
        'total_generations': total_generations,
        'cards_on_cloud': cards_on_cloud,
        'generated_voters_count': generated_voters_count,
        'db_connected': db_connected,
        'mongodb_size_mb': '...',
        'cloudinary_credits': '...',
        'sms_balance': '...',
        'total_referrals': total_referrals,
        'pending_volunteers': pending_volunteers,
        'confirmed_volunteers': confirmed_volunteers,
        'pending_booth_agents': pending_booth_agents,
        'confirmed_booth_agents': confirmed_booth_agents,
    }

    # P2: Cache in Redis for 60s
    _cache_set(REDIS_DASHBOARD_KEY, result, REDIS_DASHBOARD_TTL)
    return result


# ══════════════════════════════════════════════════════════════════
#  PUBLIC USER ROUTES  (/)
# ══════════════════════════════════════════════════════════════════

@app.route('/favicon.ico')
def favicon():
    from flask import send_from_directory
    return send_from_directory(app.static_folder, 'favicon.jpg', mimetype='image/jpeg')


@app.route('/cronjob')
def cronjob():
    return 'OK', 200, {'Content-Type': 'text/plain'}


@app.route('/')
def user_home():
    resp = app.make_response(render_template('user/chatbot.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


# ── Chatbot API Endpoints ────────────────────────────────────────

@app.route('/api/chat/send-otp', methods=['POST'])
@limiter.limit("3 per 5 minutes")  # PHASE 1: Use Redis-based rate limiter
def chat_send_otp():
    """Generate and send OTP to mobile number via 2Factor.in - rate-limited."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    
    # Validate mobile number
    valid, result = validate_mobile(mobile)
    if not valid:
        return jsonify({'success': False, 'message': result}), 400
    mobile = result

    # Rate limit: max 1 OTP per mobile per 60 seconds
    existing = otp_col.find_one({'mobile': mobile})
    if existing:
        try:
            created = datetime.fromisoformat(existing['created_at'])
            elapsed = (datetime.now(timezone.utc) - created).total_seconds()
            if elapsed < 60:
                wait = int(60 - elapsed)
                return jsonify({'success': False, 'message': f'Please wait {wait}s before requesting another OTP.'}), 429
        except Exception:
            pass

    # SECURITY FIX: Use cryptographically secure random for OTP
    otp = str(secrets.randbelow(900000) + 100000)

    # Store OTP in DB
    otp_col.update_one(
        {'mobile': mobile},
        {'$set': {
            'otp': otp,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'verified': False,
        }},
        upsert=True
    )

    # PHASE 1: Send OTP via 2Factor.in with circuit breaker
    otp_sent = False
    sms_api_key = os.getenv('SMS_API_KEY', '')
    if sms_api_key:
        try:
            # Use circuit breaker for SMS API
            @sms_breaker
            def send_sms():
                resp = http_requests.get(
                    f'https://2factor.in/API/V1/{sms_api_key}/SMS/{mobile}/{otp}',
                    timeout=15
                )
                if resp.status_code == 200 and resp.json().get('Status') == 'Success':
                    return True
                return False
            
            otp_sent = send_sms()
            if otp_sent:
                # SECURITY FIX: Mask mobile number in logs
                logger.info(f"OTP call sent to {mobile[:2]}****{mobile[-2:]}")
        except Exception as e:
            # SECURITY FIX: Mask mobile number in logs
            logger.warning(f"OTP send failed for {mobile[:2]}****{mobile[-2:]}: {e}")
            # Circuit breaker may be open
            if 'CircuitBreakerError' in str(type(e).__name__):
                logger.warning("SMS API circuit breaker is OPEN - service unavailable")

    if not otp_sent:
        # SECURITY FIX: Mask mobile number in logs
        logger.warning(f"OTP not sent for {mobile[:2]}****{mobile[-2:]}")

    return jsonify({'success': otp_sent})


@app.route('/api/chat/verify-otp', methods=['POST'])
@rate_limit(max_requests=5, window_seconds=300)  # 5 attempts per 5 minutes
def chat_verify_otp():
    """Verify OTP for mobile number."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    otp = data.get('otp', '').strip()

    # Validate inputs
    valid_mobile, mobile_result = validate_mobile(mobile)
    if not valid_mobile:
        return jsonify({'success': False, 'message': mobile_result}), 400
    mobile = mobile_result
    
    valid_otp, otp_result = validate_otp(otp)
    if not valid_otp:
        return jsonify({'success': False, 'message': otp_result}), 400
    otp = otp_result

    doc = otp_col.find_one({'mobile': mobile})
    if not doc or doc.get('otp') != otp:
        return jsonify({'success': False, 'message': 'Invalid OTP'}), 400

    # Check expiry (5 minutes)
    try:
        created = datetime.fromisoformat(doc['created_at'])
        if (datetime.now(timezone.utc) - created).total_seconds() > 300:
            return jsonify({'success': False, 'message': 'OTP expired. Please request a new one.'}), 400
    except Exception:
        pass

    # Mark OTP as verified (but don't mark mobile as verified yet - that happens after card generation)
    otp_col.update_one({'mobile': mobile}, {'$set': {'verified': True}})
    
    # SECURITY FIX: Store verified mobile in session for authorization checks
    session['verified_mobile'] = mobile
    session.permanent = True

    # Check if this mobile already has a linked card
    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1, 'card_url': 1, 'name': 1})
    if stat and stat.get('card_url'):
        # For existing users, fetch voter name from gen_voters_col
        gen_doc = gen_voters_col.find_one({'mobile': mobile}, {'name': 1, 'photo_url': 1})
        return jsonify({
            'success': True,
            'has_card': True,
            'epic_no': stat.get('epic_no', ''),
            'card_url': stat.get('card_url', ''),
            'voter_name': (gen_doc.get('name', '') if gen_doc else '') or stat.get('name', ''),
            'photo_url': (gen_doc.get('photo_url', '') if gen_doc else '')
        })

    return jsonify({'success': True, 'has_card': False})


@app.route('/api/chat/check-mobile', methods=['POST'])
def chat_check_mobile():
    """Check if a mobile number already has a linked card (and thus a PIN)."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1, 'card_url': 1, 'secret_pin': 1})
    if stat and stat.get('card_url'):
        has_pin = bool(stat.get('secret_pin'))
        return jsonify({'success': True, 'has_card': True, 'has_pin': has_pin})
    return jsonify({'success': True, 'has_card': False, 'has_pin': False})


@app.route('/api/chat/verify-pin', methods=['POST'])
@rate_limit(max_requests=5, window_seconds=300)  # 5 attempts per 5 minutes
def chat_verify_pin():
    """Verify the 4-digit secret PIN for a returning user."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    pin = data.get('pin', '').strip()

    # Validate inputs
    valid_mobile, mobile_result = validate_mobile(mobile)
    if not valid_mobile:
        return jsonify({'success': False, 'message': mobile_result}), 400
    mobile = mobile_result
    
    valid_pin, pin_result = validate_pin(pin)
    if not valid_pin:
        return jsonify({'success': False, 'message': pin_result}), 400
    pin = pin_result

    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1, 'card_url': 1, 'secret_pin': 1})
    if not stat or not stat.get('secret_pin'):
        return jsonify({'success': False, 'message': 'No PIN found for this mobile.'}), 404

    # SECURITY: Verify hashed PIN
    if not verify_pin(pin, stat['secret_pin']):
        return jsonify({'success': False, 'message': 'Invalid PIN. Please try again.'}), 400

    gen_doc = gen_voters_col.find_one({'mobile': mobile}, {'name': 1, 'photo_url': 1})
    return jsonify({
        'success': True,
        'has_card': True,
        'epic_no': stat.get('epic_no', ''),
        'card_url': stat.get('card_url', ''),
        'voter_name': (gen_doc.get('name', '') if gen_doc else ''),
        'photo_url': (gen_doc.get('photo_url', '') if gen_doc else '')
    })


@app.route('/api/chat/forgot-pin', methods=['POST'])
def chat_forgot_pin():
    """Send voice OTP for PIN reset."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    # Verify this mobile actually has a card
    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1})
    if not stat:
        return jsonify({'success': False, 'message': 'No account found for this mobile.'}), 404

    # Rate limit: max 1 OTP per mobile per 60 seconds
    existing = otp_col.find_one({'mobile': mobile})
    if existing:
        try:
            created = datetime.fromisoformat(existing['created_at'])
            elapsed = (datetime.now(timezone.utc) - created).total_seconds()
            if elapsed < 60:
                wait = int(60 - elapsed)
                return jsonify({'success': False, 'message': f'Please wait {wait}s before requesting another OTP.'}), 429
        except Exception:
            pass

    # SECURITY FIX: Use cryptographically secure random for OTP
    otp = str(secrets.randbelow(900000) + 100000)
    otp_col.update_one(
        {'mobile': mobile},
        {'$set': {
            'otp': otp,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'verified': False,
            'purpose': 'pin_reset',
        }},
        upsert=True
    )

    otp_sent = False
    sms_api_key = os.getenv('SMS_API_KEY', '')
    if sms_api_key:
        try:
            resp = http_requests.get(
                f'https://2factor.in/API/V1/{sms_api_key}/SMS/{mobile}/{otp}',
                timeout=15
            )
            if resp.status_code == 200 and resp.json().get('Status') == 'Success':
                otp_sent = True
                logger.info(f"PIN reset OTP sent to {mobile}")
        except Exception as e:
            logger.warning(f"PIN reset OTP send failed for {mobile}: {e}")

    return jsonify({'success': otp_sent})


@app.route('/api/chat/reset-pin', methods=['POST'])
@rate_limit(max_requests=3, window_seconds=300)
def chat_reset_pin():
    """Verify OTP and save user-chosen new PIN."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    otp = data.get('otp', '').strip()

    # Validate inputs
    valid_mobile, mobile_result = validate_mobile(mobile)
    if not valid_mobile:
        return jsonify({'success': False, 'message': mobile_result}), 400
    mobile = mobile_result
    
    valid_otp, otp_result = validate_otp(otp)
    if not valid_otp:
        return jsonify({'success': False, 'message': otp_result}), 400
    otp = otp_result

    doc = otp_col.find_one({'mobile': mobile})
    if not doc or doc.get('otp') != otp:
        return jsonify({'success': False, 'message': 'Invalid OTP'}), 400

    # Check expiry (5 minutes)
    try:
        created = datetime.fromisoformat(doc['created_at'])
        if (datetime.now(timezone.utc) - created).total_seconds() > 300:
            return jsonify({'success': False, 'message': 'OTP expired. Please request a new one.'}), 400
    except Exception:
        pass

    # Accept user-chosen PIN
    new_pin = data.get('new_pin', '').strip()
    valid_pin, pin_result = validate_pin(new_pin)
    if not valid_pin:
        return jsonify({'success': False, 'message': pin_result}), 400
    new_pin = pin_result

    # SECURITY: Hash PIN before saving
    hashed_pin = hash_pin(new_pin)
    stats_col.update_one({'auth_mobile': mobile}, {'$set': {'secret_pin': hashed_pin}})
    gen_voters_col.update_one({'mobile': mobile}, {'$set': {'secret_pin': hashed_pin}})
    # Clean up OTP
    otp_col.delete_one({'mobile': mobile})

    # Get card info to return
    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1, 'card_url': 1})
    gen_doc = gen_voters_col.find_one({'mobile': mobile}, {'name': 1, 'photo_url': 1})

    logger.info(f"PIN reset successful for {mobile}")

    return jsonify({
        'success': True,
        'has_card': True,
        'epic_no': stat.get('epic_no', '') if stat else '',
        'card_url': stat.get('card_url', '') if stat else '',
        'voter_name': (gen_doc.get('name', '') if gen_doc else ''),
        'photo_url': (gen_doc.get('photo_url', '') if gen_doc else '')
    })


@app.route('/api/chat/set-pin', methods=['POST'])
@rate_limit(max_requests=3, window_seconds=300)
def chat_set_pin():
    """Set the 4-digit PIN for a user who just registered (card exists, no PIN yet)."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    pin = data.get('pin', '').strip()

    # Validate inputs
    valid_mobile, mobile_result = validate_mobile(mobile)
    if not valid_mobile:
        return jsonify({'success': False, 'message': mobile_result}), 400
    mobile = mobile_result
    
    valid_pin, pin_result = validate_pin(pin)
    if not valid_pin:
        return jsonify({'success': False, 'message': pin_result}), 400
    pin = pin_result

    # Verify this mobile has a card
    stat = stats_col.find_one({'auth_mobile': mobile}, {'epic_no': 1, 'card_url': 1})
    if not stat or not stat.get('card_url'):
        return jsonify({'success': False, 'message': 'No card found for this mobile.'}), 404

    # SECURITY: Hash PIN before saving
    hashed_pin = hash_pin(pin)
    stats_col.update_one({'auth_mobile': mobile}, {'$set': {'secret_pin': hashed_pin}})
    gen_voters_col.update_one({'mobile': mobile}, {'$set': {'secret_pin': hashed_pin}})

    logger.info(f"PIN set for {mobile}")
    return jsonify({'success': True})


@app.route('/api/chat/verify-forgot-otp', methods=['POST'])
def chat_verify_forgot_otp():
    """Verify the OTP sent for forgot-PIN flow (does NOT delete OTP or reset PIN)."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    otp = data.get('otp', '').strip()

    if not mobile or not otp:
        return jsonify({'success': False, 'message': 'Mobile and OTP required'}), 400

    doc = otp_col.find_one({'mobile': mobile})
    if not doc or doc.get('otp') != otp:
        return jsonify({'success': False, 'message': 'Invalid OTP'}), 400

    try:
        created = datetime.fromisoformat(doc['created_at'])
        if (datetime.now(timezone.utc) - created).total_seconds() > 300:
            return jsonify({'success': False, 'message': 'OTP expired. Please request a new one.'}), 400
    except Exception:
        pass

    return jsonify({'success': True})


@app.route('/api/chat/validate-epic', methods=['POST'])
@rate_limit(max_requests=10, window_seconds=60)
def chat_validate_epic():
    """Validate EPIC number and return voter details."""
    data = request.get_json()
    epic_no = data.get('epic_no', '').strip().upper()
    
    # Validate EPIC format
    valid, result = validate_epic(epic_no)
    if not valid:
        return jsonify({'success': False, 'message': result}), 400
    epic_no = result

    voter = find_voter_by_epic(epic_no)
    if not voter:
        return jsonify({'success': False, 'message': 'EPIC Number not found. Please check and try again.'}), 404

    display_fields = {}
    for key, val in voter.items():
        if key.startswith('_'):
            continue
        display_fields[key] = str(val) if val else ''

    return jsonify({'success': True, 'voter': display_fields})


@app.route('/api/chat/generate-card', methods=['POST'])
@limiter.limit("5 per 5 minutes")  # PHASE 1: Use Redis-based rate limiter
def chat_generate_card():
    """Upload photo and generate ID card via chatbot - ASYNC with Celery."""
    epic_no = request.form.get('epic_no', '').strip().upper()
    mobile_num = request.form.get('mobile', '').strip()

    # Validate EPIC
    valid_epic, epic_result = validate_epic(epic_no)
    if not valid_epic:
        return jsonify({'success': False, 'message': epic_result}), 400
    epic_no = epic_result

    voter = find_voter_by_epic(epic_no)
    if not voter:
        return jsonify({'success': False, 'message': 'EPIC Number not found.'}), 404

    # Photo is required
    if 'photo' not in request.files or not request.files['photo'].filename:
        return jsonify({'success': False, 'message': 'Photo is required.'}), 400

    file = request.files['photo']
    
    # SECURITY: Validate file upload
    valid_file, file_error = validate_file_upload(file, ALLOWED_IMG, max_size_mb=10)
    if not valid_file:
        return jsonify({'success': False, 'message': file_error}), 400

    try:
        # FACE DETECTION: Validate photo contains a clear face
        is_valid_face, face_message, photo_image = validate_photo_for_id_card(file.stream)
        
        if not is_valid_face:
            logger.warning(f"Face detection failed for {epic_no}: {face_message}")
            return jsonify({
                'success': False, 
                'message': face_message,
                'error_type': 'face_detection_failed'
            }), 400
        
        logger.info(f"Face detected successfully for {epic_no}")
        
        # Convert validated photo to base64 for Celery task
        photo_buffer = io.BytesIO()
        photo_image.save(photo_buffer, format='JPEG', quality=95)
        photo_buffer.seek(0)
        
        import base64
        photo_base64 = base64.b64encode(photo_buffer.getvalue()).decode('utf-8')
        
        # Generate unique PTC code
        ptc_code = generate_ptc_code()
        
        # Get referral info
        ref_ptc = request.form.get('ref_ptc', '').strip()
        ref_rid = request.form.get('ref_rid', '').strip()
        
        # PHASE 1: Queue card generation as async Celery task
        task = generate_card_async.delay(
            epic_no=epic_no,
            mobile=mobile_num,
            photo_base64=photo_base64,
            ptc_code=ptc_code,
            referred_by_ptc=ref_ptc,
            referred_by_referral_id=ref_rid,
            secret_pin=''
        )
        
        logger.info(f"Card generation queued for {epic_no}, task_id: {task.id}")
        
        # Return immediately with job ID
        return jsonify({
            'success': True,
            'job_id': task.id,
            'status': 'processing',
            'message': 'Card generation started. Please check status.',
            'epic_no': epic_no
        })
        
    except Exception as e:
        logger.error(f"Card generation queue error for {epic_no}: {e}")
        return jsonify({'success': False, 'message': 'Failed to queue card generation. Please try again.'}), 500


@app.route('/api/chat/card-status/<job_id>')
def check_card_status(job_id):
    """Check status of async card generation job."""
    try:
        from celery.result import AsyncResult
        task = AsyncResult(job_id, app=celery)
        
        if task.state == 'PENDING':
            response = {
                'status': 'pending',
                'message': 'Job is waiting to be processed'
            }
        elif task.state == 'PROCESSING':
            response = {
                'status': 'processing',
                'message': task.info.get('status', 'Processing...') if task.info else 'Processing...'
            }
        elif task.state == 'SUCCESS':
            result = task.result
            response = {
                'status': 'completed',
                'success': result.get('success', False),
                'card_url': result.get('card_url', ''),
                'photo_url': result.get('photo_url', ''),
                'epic_no': result.get('epic_no', ''),
                'voter_name': result.get('voter_name', ''),
                'message': result.get('message', 'Card generated successfully')
            }
            
            # Mark mobile as verified after successful generation
            if result.get('success') and result.get('epic_no'):
                mobile = session.get('verified_mobile')
                if mobile:
                    verified_mobiles_col.update_one(
                        {'mobile': mobile},
                        {'$set': {
                            'mobile': mobile,
                            'epic_no': result['epic_no'],
                            'verified_at': datetime.now(timezone.utc).isoformat()
                        }},
                        upsert=True
                    )
        elif task.state == 'FAILURE':
            response = {
                'status': 'failed',
                'message': str(task.info) if task.info else 'Card generation failed'
            }
        else:
            response = {
                'status': task.state.lower(),
                'message': f'Job status: {task.state}'
            }
        
        return jsonify(response)
        
    except Exception as e:
        logger.error(f"Error checking job status {job_id}: {e}")
        return jsonify({
            'status': 'error',
            'message': 'Failed to check job status'
        }), 500


@app.route('/card/<epic_no>')
def user_card_page(epic_no):
    """GET page showing generated card - safe to refresh (no re-generation)."""
    epic_no = epic_no.strip().upper()
    voter = find_voter_by_epic(epic_no)
    if not voter:
        flash('Voter not found.', 'danger')
        return redirect(url_for('user_home'))

    card_url = get_voter_card_url(epic_no)
    if not card_url:
        flash('Card not generated yet. Please generate first.', 'warning')
        return redirect(url_for('user_home'))

    gen_count = get_voter_gen_count(epic_no)
    return render_template('user/card.html',
                           epic_no=epic_no,
                           voter=voter,
                           gen_count=gen_count,
                           card_url=card_url)


@app.route('/mycard/<epic_no>')
def user_serve_card(epic_no):
    """Redirect to the Cloudinary-hosted card image - REQUIRES AUTHORIZATION."""
    # SECURITY FIX: Verify user owns this card
    mobile = session.get('verified_mobile')
    if not mobile:
        return jsonify({'error': 'Unauthorized. Please verify your mobile number first.'}), 401
    
    # Verify this mobile generated this card
    epic_no = epic_no.strip().upper()
    gen_doc = gen_voters_col.find_one({'epic_no': epic_no, 'mobile': mobile})
    if not gen_doc:
        return jsonify({'error': 'Forbidden. You do not have access to this card.'}), 403
    
    card_url = get_voter_card_url(epic_no)
    if card_url:
        return redirect(card_url)
    return jsonify({'error': 'Card not found.'}), 404


@app.route('/mycard/<epic_no>/download')
def user_download_card(epic_no):
    """Redirect to Cloudinary card with download flag - REQUIRES AUTHORIZATION."""
    # SECURITY FIX: Verify user owns this card
    mobile = session.get('verified_mobile')
    if not mobile:
        return jsonify({'error': 'Unauthorized. Please verify your mobile number first.'}), 401
    
    # Verify this mobile generated this card
    epic_no = epic_no.strip().upper()
    gen_doc = gen_voters_col.find_one({'epic_no': epic_no, 'mobile': mobile})
    if not gen_doc:
        return jsonify({'error': 'Forbidden. You do not have access to this card.'}), 403
    
    card_url = get_voter_card_url(epic_no)
    if card_url:
        # Add Cloudinary fl_attachment transformation for download
        if '/upload/' in card_url:
            dl_url = card_url.replace('/upload/', f'/upload/fl_attachment:{epic_no}_VoterID/')
        else:
            dl_url = card_url
        return redirect(dl_url)
    return jsonify({'error': 'Card not found.'}), 404


@app.route('/verify/<epic_no>')
def verify_voter(epic_no):
    """Public verification page - opened when QR code is scanned on mobile."""
    epic_no = epic_no.strip().upper()
    voter = find_voter_by_epic(epic_no)
    if not voter:
        flash('Voter ID not found in database.', 'danger')
        return redirect(url_for('user_home'))

    # Attach stats (exclude secret_pin — never expose in QR/verify page)
    s = stats_col.find_one({'epic_no': epic_no}, {'_id': 0, 'secret_pin': 0}) or {}
    voter['gen_count'] = s.get('count', 0)
    voter['last_generated'] = s.get('last_generated', '')
    voter['photo_url'] = s.get('photo_url', '')
    voter['card_url'] = s.get('card_url', '')
    
    # SECURITY FIX: Don't expose full mobile number publicly - show last 4 digits only
    mobile = s.get('auth_mobile', '')
    if mobile and len(mobile) >= 4:
        voter['auth_mobile_masked'] = f"****{mobile[-4:]}"
    else:
        voter['auth_mobile_masked'] = ''

    # Attach PTC code from generated voters DB
    gen_doc = gen_voters_col.find_one({'epic_no': epic_no}, {'ptc_code': 1})
    voter['ptc_code'] = gen_doc.get('ptc_code', '') if gen_doc else ''

    # Separate core fields from extra fields
    core_keys = {'epic_no', 'name', 'assembly', 'district', 'gen_count',
                 'last_generated', 'photo_url', 'card_url', 'serial_number',
                 'verify_url', 'auth_mobile', 'ptc_code'}
    extra_fields = {k: v for k, v in voter.items() if k not in core_keys and v}

    return render_template('user/verify.html',
                           voter=voter,
                           extra_fields=extra_fields)


# ── Referral Landing Page ────────────────────────────────────────

@app.route('/refer/<ptc_code>/<referral_id>')
def referral_landing(ptc_code, referral_id):
    """Serve OG-tagged page for WhatsApp preview, then redirect to chatbot."""
    voter = gen_voters_col.find_one({'ptc_code': ptc_code, 'referral_id': referral_id})
    if not voter:
        flash('Invalid referral link.', 'danger')
        return redirect(url_for('user_home'))

    referrer_name = voter.get('name', 'A PuratchiThaai Member')
    redirect_url = url_for('user_home') + f'?ref={ptc_code}&rid={referral_id}'
    banner_url = f"{config.BASE_URL}/static/banner.jpg"

    html = f"""<!DOCTYPE html>
<html><head>
<meta charset="UTF-8">
<meta property="og:title" content="PuratchiThaai — Become a Member!">
<meta property="og:description" content="{referrer_name} invites you to join PuratchiThaai! Generate your free Digital Member ID Card now and become a proud member.">
<meta property="og:image" content="{banner_url}">
<meta property="og:image:width" content="1200">
<meta property="og:image:height" content="630">
<meta property="og:type" content="website">
<meta property="og:url" content="{config.BASE_URL}/refer/{ptc_code}/{referral_id}">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="PuratchiThaai — Become a Member!">
<meta name="twitter:description" content="{referrer_name} invites you to join PuratchiThaai! Generate your free Digital Member ID Card now.">
<meta name="twitter:image" content="{banner_url}">
<meta http-equiv="refresh" content="0;url={redirect_url}">
<title>PuratchiThaai — Join Now!</title>
</head><body style="font-family:sans-serif;text-align:center;padding:40px;">
<p>Redirecting to PuratchiThaai...</p>
<script>window.location.href="{redirect_url}";</script>
</body></html>"""
    return html


# ── New Chatbot API Endpoints ────────────────────────────────────

@app.route('/api/chat/profile', methods=['POST'])
def chat_profile():
    """Return voter profile details for sidebar profile view."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    gen_doc = gen_voters_col.find_one({'mobile': mobile}, {'_id': 0, 'secret_pin': 0})
    if not gen_doc:
        return jsonify({'success': False, 'message': 'Profile not found.'}), 404

    # Also fetch original voter data for additional fields
    voter_doc = None
    epic = gen_doc.get('epic_no', '')
    if epic:
        voter_doc = voters_col.find_one({'epic_no': epic}, {'_id': 0})

    # Merge: gen_doc fields take priority, but include extra voter_doc fields
    profile = {}
    if voter_doc:
        profile.update(voter_doc)
    profile.update({k: v for k, v in gen_doc.items() if v})

    # Remove sensitive fields
    profile.pop('secret_pin', None)
    profile.pop('_id', None)

    return jsonify({'success': True, 'profile': profile})


@app.route('/api/chat/booth', methods=['POST'])
def chat_booth():
    """Return booth / polling station info for navigation.
    Auto-detects lat,long and address from any voter field."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    gen_doc = gen_voters_col.find_one({'mobile': mobile}, {'_id': 0})
    if not gen_doc:
        return jsonify({'success': False, 'message': 'Profile not found.'}), 404

    epic = gen_doc.get('epic_no', '')
    voter_doc = voters_col.find_one({'epic_no': epic}, {'_id': 0}) if epic else None

    # Merge to get all fields
    merged = {}
    if voter_doc:
        merged.update(voter_doc)
    merged.update({k: v for k, v in gen_doc.items() if v})

    # Skip these keys when scanning for booth data
    skip_keys = {'_id', 'epic_no', 'name', 'mobile', 'photo_url', 'card_url',
                 'ptc_code', 'secret_pin', 'referral_id', 'referred_by_ptc',
                 'referred_by_referral_id', 'generated_at', 'referred_members_count'}

    # Regex: lat,long like 13.2296228,79.6741222
    latlong_re = re.compile(
        r'^[-+]?([1-8]?\d(\.\d+)?|90(\.0+)?)\s*,\s*[-+]?(180(\.0+)?|(1[0-7]\d|\d{1,2})(\.\d+)?)$'
    )
    # Address heuristic: contains a 6-digit pincode, or is long text (>30 chars)
    # with words like school, room, road, street, building, village, panchayat, ward, etc.
    addr_keywords = re.compile(
        r'school|room|road|street|building|village|panchayat|ward|nagar|colony|'
        r'block|floor|hall|office|temple|church|mosque|community|union|elementary|'
        r'middle|higher|secondary|primary|facing|north|south|east|west',
        re.IGNORECASE
    )

    detected_latlong = None
    detected_address = None
    latlong_key = None
    address_key = None

    for key, val in merged.items():
        if key in skip_keys or not val:
            continue
        s = str(val).strip()
        if not s:
            continue

        # Check for lat,long pattern
        if not detected_latlong and latlong_re.match(s):
            detected_latlong = s
            latlong_key = key
            continue

        # Check for address-like value
        if not detected_address:
            has_pincode = bool(re.search(r'\b\d{6}\b', s))
            has_keywords = bool(addr_keywords.search(s))
            if (has_pincode or has_keywords) and len(s) > 20:
                detected_address = s
                address_key = key

    # Also check explicit fields from column mapping
    if not detected_address:
        for f in ('polling_station', 'booth_address'):
            if merged.get(f):
                detected_address = merged[f]
                address_key = f
                break

    booth = {
        'latlong': detected_latlong or '',
        'address': detected_address or '',
        'latlong_field': latlong_key or '',
        'address_field': address_key or '',
        'assembly': merged.get('assembly', ''),
        'district': merged.get('district', ''),
        'part_no': merged.get('part_no', ''),
        'polling_station': merged.get('polling_station', ''),
    }

    has_data = bool(detected_latlong or detected_address)
    return jsonify({'success': True, 'booth': booth, 'has_booth_data': has_data})


@app.route('/api/chat/get-referral-link', methods=['POST'])
def chat_get_referral_link():
    """Get or create unique referral link for a voter."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    voter = gen_voters_col.find_one({'mobile': mobile})
    if not voter or not voter.get('ptc_code'):
        return jsonify({'success': False, 'message': 'Voter not found.'}), 404

    result = get_or_create_referral(voter['ptc_code'])
    if not result:
        return jsonify({'success': False, 'message': 'Could not generate referral link.'}), 500

    return jsonify({
        'success': True,
        'referral_id': result['referral_id'],
        'referral_link': result['referral_link'],
        'ptc_code': voter['ptc_code']
    })


@app.route('/api/chat/my-members', methods=['POST'])
def chat_my_members():
    """Get list of voters referred by this voter."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    voter = gen_voters_col.find_one({'mobile': mobile})
    if not voter or not voter.get('ptc_code'):
        return jsonify({'success': False, 'message': 'Voter not found.'}), 404

    ptc_code = voter['ptc_code']
    members = list(gen_voters_col.find(
        {'referred_by_ptc': ptc_code},
        {'_id': 0, 'name': 1, 'epic_no': 1, 'assembly': 1, 'district': 1,
         'ptc_code': 1, 'generated_at': 1, 'mobile': 1}
    ).sort('generated_at', -1))

    return jsonify({
        'success': True,
        'referrer_name': voter.get('name', ''),
        'referrer_ptc': ptc_code,
        'total_referred': len(members),
        'members': members
    })


@app.route('/api/chat/request-volunteer', methods=['POST'])
def chat_request_volunteer():
    """Submit a volunteer request."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    voter = gen_voters_col.find_one({'mobile': mobile})
    if not voter or not voter.get('ptc_code'):
        return jsonify({'success': False, 'message': 'Voter not found.'}), 404

    existing = volunteer_requests_col.find_one({'ptc_code': voter['ptc_code']})
    if existing:
        return jsonify({
            'success': True,
            'already_requested': True,
            'status': existing.get('status', 'pending'),
            'message': 'You have already submitted a volunteer request.'
        })

    doc = {
        'ptc_code': voter['ptc_code'],
        'epic_no': voter.get('epic_no', ''),
        'name': voter.get('name', ''),
        'mobile': mobile,
        'assembly': voter.get('assembly', ''),
        'district': voter.get('district', ''),
        'photo_url': voter.get('photo_url', ''),
        'status': 'pending',
        'requested_at': datetime.now(timezone.utc).isoformat(),
        'reviewed_at': None,
        'reviewed_by': None,
    }
    volunteer_requests_col.insert_one(doc)
    return jsonify({'success': True, 'already_requested': False, 'message': 'Volunteer request submitted.'})


@app.route('/api/chat/request-booth-agent', methods=['POST'])
def chat_request_booth_agent():
    """Submit a booth agent request."""
    data = request.get_json()
    mobile = data.get('mobile', '').strip()
    if not mobile or len(mobile) != 10:
        return jsonify({'success': False, 'message': 'Invalid mobile number'}), 400

    voter = gen_voters_col.find_one({'mobile': mobile})
    if not voter or not voter.get('ptc_code'):
        return jsonify({'success': False, 'message': 'Voter not found.'}), 404

    existing = booth_agent_requests_col.find_one({'ptc_code': voter['ptc_code']})
    if existing:
        return jsonify({
            'success': True,
            'already_requested': True,
            'status': existing.get('status', 'pending'),
            'message': 'You have already submitted a booth agent request.'
        })

    doc = {
        'ptc_code': voter['ptc_code'],
        'epic_no': voter.get('epic_no', ''),
        'name': voter.get('name', ''),
        'mobile': mobile,
        'assembly': voter.get('assembly', ''),
        'district': voter.get('district', ''),
        'photo_url': voter.get('photo_url', ''),
        'status': 'pending',
        'requested_at': datetime.now(timezone.utc).isoformat(),
        'reviewed_at': None,
        'reviewed_by': None,
    }
    booth_agent_requests_col.insert_one(doc)
    return jsonify({'success': True, 'already_requested': False, 'message': 'Booth agent request submitted.'})


@app.route('/api/whatsapp-channel')
def api_whatsapp_channel():
    """Return WhatsApp channel URL."""
    url = config.WHATSAPP_CHANNEL_URL
    return jsonify({'url': url})


# ══════════════════════════════════════════════════════════════════
#  ADMIN BLUEPRINT  (/admin)
# ══════════════════════════════════════════════════════════════════

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


# ── Admin Authentication ─────────────────────────────────────────
@admin_bp.before_request
def require_admin_login():
    """Guard all /admin routes - redirect to login if not authenticated."""
    if request.endpoint == 'admin.login':
        return  # allow access to the login page itself
    if not session.get('admin_logged_in'):
        flash('Please log in to access the admin panel.', 'warning')
        return redirect(url_for('admin.login', next=request.url))


@admin_bp.route('/login', methods=['GET', 'POST'])
@rate_limit(max_requests=5, window_seconds=300)  # 5 login attempts per 5 minutes
def login():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin.dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        ip = request.remote_addr or 'unknown'
        
        # SECURITY: Check if IP is locked out
        is_locked, retry_after = login_tracker.is_locked(ip)
        if is_locked:
            flash(f'Too many failed attempts. Try again in {retry_after} seconds.', 'danger')
            return render_template('admin/login.html')
        
        # Verify credentials
        if username == config.ADMIN_USERNAME and password == config.ADMIN_PASSWORD:
            # SECURITY FIX: Regenerate session on login to prevent session fixation
            session.clear()
            session['admin_logged_in'] = True
            session.permanent = True
            login_tracker.reset(ip)  # Reset on successful login
            flash('Welcome back, Admin!', 'success')
            next_url = request.args.get('next') or url_for('admin.dashboard')
            return redirect(next_url)
        else:
            # SECURITY: Record failed attempt
            login_tracker.record_attempt(ip, username, success=False)
            flash('Invalid username or password.', 'danger')
    
    return render_template('admin/login.html')


@admin_bp.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    flash('Logged out successfully.', 'info')
    return redirect(url_for('admin.login'))


@admin_bp.route('/')
def dashboard():
    stats = get_dashboard_stats()
    return render_template('admin/dashboard.html', stats=stats)


@admin_bp.route('/voters')
@rate_limit(max_requests=30, window_seconds=60)
def voters_list():
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    # Only fetch lightweight data for the template shell
    # Actual voter data is loaded via AJAX from /api/voters
    assemblies, districts = _get_cached_dropdowns(voters_col, REDIS_DROPDOWN_VOTERS_KEY)
    try:
        total = voters_col.estimated_document_count()
    except Exception:
        total = 0

    return render_template('admin/voters.html',
                           voters=[], page=1,
                           total_pages=1, total=total,
                           per_page=per_page, search='',
                           assemblies=assemblies, districts=districts,
                           filter_assembly='',
                           filter_district='')


def _run_import_thread(file_path: str, csv_bytes: bytes, ext: str, import_mode: str):
    """Background thread: parse file and stream-insert into MongoDB."""
    global import_status
    try:
        import_status['phase'] = 'parsing'

        if ext == 'csv':
            voter_iter = _iter_csv_bytes(csv_bytes)
        else:
            voter_iter = _iter_xlsx(file_path)

        import_status['phase'] = 'inserting'

        if import_mode == 'replace':
            count = _stream_replace(voter_iter, import_status)
            import_status['message'] = f'Data replaced successfully! {count} unique voter records stored.'
        else:
            existing_count = voters_col.count_documents({})
            count = _stream_upsert(voter_iter, import_status)
            new_total = voters_col.count_documents({})
            new_added = new_total - existing_count
            updated = count - new_added if count > new_added else 0
            import_status['message'] = (
                f'Import merged! {import_status["processed"]} records processed - '
                f'{new_added} new added, {updated} updated, '
                f'{new_total} total in database.'
            )

        import_status['phase'] = 'done'

    except Exception as e:
        import_status['phase'] = 'error'
        import_status['error'] = str(e)
    finally:
        import_status['running'] = False


@admin_bp.route('/import', methods=['GET', 'POST'])
def import_xlsx():
    global import_status

    if request.method == 'POST':
        # Prevent concurrent imports
        if import_status.get('running'):
            flash('An import is already in progress. Please wait.', 'warning')
            return redirect(url_for('admin.import_xlsx'))

        if 'file' not in request.files:
            flash('No file selected.', 'danger')
            return redirect(url_for('admin.import_xlsx'))

        file = request.files['file']
        if not file or not file.filename:
            flash('No file selected.', 'danger')
            return redirect(url_for('admin.import_xlsx'))

        if not allowed_file(file.filename, ALLOWED_DATA):
            flash('Only .xlsx, .xls, and .csv files are accepted.', 'danger')
            return redirect(url_for('admin.import_xlsx'))

        ext = file.filename.rsplit('.', 1)[1].lower()
        import_mode = request.form.get('import_mode', 'merge')

        # Save file / read bytes BEFORE launching thread
        csv_bytes = None
        file_path = None
        try:
            if ext == 'csv':
                csv_bytes = file.stream.read()
            else:
                os.makedirs(config.DATA_DIR, exist_ok=True)
                file.save(config.VOTERS_XLSX)
                file_path = config.VOTERS_XLSX
        except Exception as e:
            flash(f'Could not read file: {e}', 'danger')
            return redirect(url_for('admin.import_xlsx'))

        # Reset status and launch background thread
        import_status.update({
            'running': True,
            'phase': 'parsing',
            'processed': 0,
            'inserted': 0,
            'total': 0,
            'message': '',
            'error': '',
        })

        t = threading.Thread(
            target=_run_import_thread,
            args=(file_path, csv_bytes, ext, import_mode),
            daemon=True
        )
        t.start()

        # Return JSON so the frontend stays on the page and polls
        return jsonify({'ok': True, 'status': 'started'})

    # GET - show current status
    try:
        voters_count = voters_col.estimated_document_count()
    except Exception:
        voters_count = 0
    return render_template('admin/import.html', voters_count=voters_count)


@admin_bp.route('/api/import-status')
def api_import_status():
    """Polling endpoint for import progress."""
    return jsonify(import_status)


@admin_bp.route('/api/stats')
def api_stats():
    return jsonify(get_dashboard_stats())


@admin_bp.route('/api/external-stats')
def api_external_stats():
    """Separate endpoint for slow external stats - loaded via AJAX."""
    return jsonify(_get_external_stats())


@admin_bp.route('/voters/<epic_no>')
def voter_detail(epic_no):
    """Show full details for a single voter."""
    voter = find_voter_by_epic(epic_no)
    if not voter:
        flash('Voter not found.', 'danger')
        return redirect(url_for('admin.voters_list'))

    # Attach stats
    s = stats_col.find_one({'epic_no': epic_no}, {'_id': 0}) or {}
    voter['gen_count'] = s.get('count', 0)
    voter['last_generated'] = s.get('last_generated', '')
    voter['photo_url'] = s.get('photo_url', '')
    voter['card_url'] = s.get('card_url', '')
    voter['auth_mobile'] = s.get('auth_mobile', '')

    # Separate core fields from extra fields for display
    core_keys = {'epic_no', 'name', 'assembly', 'district', 'gen_count',
                 'last_generated', 'photo_url', 'card_url', 'auth_mobile'}
    extra_fields = {k: v for k, v in voter.items() if k not in core_keys}

    return render_template('admin/voter_detail.html',
                           voter=voter, extra_fields=extra_fields)

@admin_bp.route('/api/voters')
@rate_limit(max_requests=30, window_seconds=60)
def api_voters():
    """JSON API: search, filter, paginate voters - cursor-based and page-based."""
    search = sanitize_search(request.args.get('search', '').strip())  # SECURITY: Sanitize search
    assembly = request.args.get('assembly', '').strip()
    district = request.args.get('district', '').strip()
    cursor = request.args.get('cursor', '').strip()
    direction = request.args.get('direction', 'next').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    # Build MongoDB query - all filtering happens in DB, not Python
    conditions = []
    if assembly:
        conditions.append({'assembly': assembly})
    if district:
        conditions.append({'district': district})
    if search:
        sf = _build_search_filter(search, ['epic_no', 'name', 'assembly', 'district'])
        if sf:
            conditions.append(sf)
    base_query = _merge_conditions(conditions)

    # Server-side count (estimated for unfiltered, count_documents for filtered)
    if base_query:
        total = voters_col.count_documents(base_query)
    else:
        total = voters_col.estimated_document_count()

    if cursor:
        # ── P2: Cursor-based pagination (avoids skip at 5.6 Cr scale) ──
        cursor_filter = _build_cursor_filter(cursor, direction, descending=False)
        query = _merge_conditions([base_query, cursor_filter]) if cursor_filter else base_query

        sort_dir = ASCENDING
        if direction == 'prev':
            sort_dir = DESCENDING
        voters = list(voters_col.find(query).sort('_id', sort_dir).limit(per_page))
        if direction == 'prev':
            voters.reverse()

        next_cursor = str(voters[-1]['_id']) if len(voters) == per_page else None
        prev_cursor = str(voters[0]['_id']) if voters else None
        for v in voters:
            v['_id'] = str(v['_id'])
    else:
        # ── Page-based (first load / fallback) ──
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        skip = (page - 1) * per_page
        voters = list(voters_col.find(base_query).sort('_id', ASCENDING).skip(skip).limit(per_page))

        next_cursor = str(voters[-1]['_id']) if len(voters) == per_page else None
        prev_cursor = None
        for v in voters:
            v['_id'] = str(v['_id'])

    # Batch-fetch stats for ONLY this page (20 records) instead of ALL stats
    epic_nos = [v['epic_no'] for v in voters if v.get('epic_no')]
    stats_docs = {s['epic_no']: s for s in stats_col.find(
        {'epic_no': {'$in': epic_nos}}, {'_id': 0}
    )} if epic_nos else {}

    for v in voters:
        s = stats_docs.get(v.get('epic_no', ''), {})
        v['gen_count'] = s.get('count', 0)
        v['last_generated'] = s.get('last_generated', '')
        v['photo_url'] = s.get('photo_url', '')
        v['card_url'] = s.get('card_url', '')
        v['auth_mobile'] = s.get('auth_mobile', '')

    response = {
        'voters': voters,
        'total': total,
        'per_page': per_page,
        'next_cursor': next_cursor,
        'prev_cursor': prev_cursor,
    }
    if cursor:
        response['has_next'] = len(voters) == per_page
        response['has_prev'] = bool(cursor)
        response['cursor_mode'] = True
    else:
        response['page'] = page
        response['total_pages'] = total_pages
        response['cursor_mode'] = False

    return jsonify(response)


# ── Generated Voters (from new DB) ──────────────────────────────

@admin_bp.route('/generated-voters')
@rate_limit(max_requests=30, window_seconds=60)
def generated_voters_list():
    """Show all voters who generated ID cards via the chatbot."""
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    # Only fetch lightweight count for the template shell
    # Actual data is loaded via AJAX from /api/generated-voters
    try:
        total = gen_voters_col.estimated_document_count()
    except Exception:
        total = 0

    return render_template('admin/generated_voters.html',
                           voters=[], page=1,
                           total_pages=1, total=total,
                           per_page=per_page, search='')


@admin_bp.route('/api/generated-voters')
@rate_limit(max_requests=30, window_seconds=60)
def api_generated_voters():
    """JSON API for generated voters list - supports cursor-based & page-based pagination."""
    search = sanitize_search(request.args.get('search', '').strip())  # SECURITY: Sanitize search
    assembly = request.args.get('assembly', '').strip()
    district = request.args.get('district', '').strip()
    cursor = request.args.get('cursor', '').strip()
    direction = request.args.get('direction', 'next').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    # Cached dropdown options (5 min TTL - distinct() is expensive at scale)
    assemblies, districts = _get_cached_dropdowns(gen_voters_col, REDIS_DROPDOWN_GEN_KEY)

    # Build base MongoDB query - all filtering in DB
    conditions = []
    if assembly:
        conditions.append({'assembly': {'$regex': f'^{re.escape(assembly)}$', '$options': 'i'}})
    if district:
        conditions.append({'district': {'$regex': f'^{re.escape(district)}$', '$options': 'i'}})
    if search:
        sf = _build_search_filter(search, ['epic_no', 'name', 'ptc_code', 'mobile', 'assembly', 'district'])
        if sf:
            conditions.append(sf)
    base_query = _merge_conditions(conditions)

    # Server-side count (estimated for unfiltered, count_documents for filtered)
    if base_query:
        total = gen_voters_col.count_documents(base_query)
    else:
        total = gen_voters_col.estimated_document_count()

    if cursor:
        # ── P2: Cursor-based pagination (no .skip - O(log n) for deep pages) ──
        cursor_filter = _build_cursor_filter(cursor, direction, descending=True)
        query = _merge_conditions([base_query, cursor_filter]) if cursor_filter else base_query

        sort_dir = DESCENDING
        if direction == 'prev':
            sort_dir = ASCENDING
        voters = list(gen_voters_col.find(query).sort('_id', sort_dir).limit(per_page))
        if direction == 'prev':
            voters.reverse()

        next_cursor = str(voters[-1]['_id']) if len(voters) == per_page else None
        prev_cursor = str(voters[0]['_id']) if voters else None
        for v in voters:
            v['_id'] = str(v['_id'])

        return jsonify({
            'voters': voters,
            'total': total,
            'per_page': per_page,
            'next_cursor': next_cursor,
            'prev_cursor': prev_cursor,
            'has_next': len(voters) == per_page,
            'has_prev': bool(cursor),
            'assemblies': assemblies,
            'districts': districts,
            'cursor_mode': True,
        })
    else:
        # ── Page-based (first load / fallback) ──
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        skip = (page - 1) * per_page
        voters = list(gen_voters_col.find(base_query).sort('_id', DESCENDING).skip(skip).limit(per_page))

        next_cursor = str(voters[-1]['_id']) if len(voters) == per_page else None
        prev_cursor = None
        for v in voters:
            v['_id'] = str(v['_id'])

        return jsonify({
            'voters': voters,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': total_pages,
            'next_cursor': next_cursor,
            'prev_cursor': prev_cursor,
            'assemblies': assemblies,
            'districts': districts,
            'cursor_mode': False,
        })


# ── Generated Voter Detail ──────────────────────────────────────

@admin_bp.route('/generated-voters/<ptc_code>')
def generated_voter_detail(ptc_code):
    """Show full details for a generated voter including referred voters."""
    voter = gen_voters_col.find_one({'ptc_code': ptc_code}, {'_id': 0})
    if not voter:
        flash('Generated voter not found.', 'danger')
        return redirect(url_for('admin.generated_voters_list'))

    # Get referred voters
    referred = list(gen_voters_col.find(
        {'referred_by_ptc': ptc_code},
        {'_id': 0, 'name': 1, 'epic_no': 1, 'ptc_code': 1, 'mobile': 1,
         'assembly': 1, 'district': 1, 'generated_at': 1, 'photo_url': 1}
    ).sort('generated_at', -1))

    return render_template('admin/generated_voter_detail.html',
                           voter=voter, referred=referred)


# ── Volunteer Requests ───────────────────────────────────────────

@admin_bp.route('/volunteer-requests')
def volunteer_requests_page():
    """Show volunteer requests."""
    return render_template('admin/volunteer_requests.html')


@admin_bp.route('/api/volunteer-requests')
def api_volunteer_requests():
    """JSON API for volunteer requests - server-side search & pagination."""
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    conditions = []
    if status:
        conditions.append({'status': status})
    if search:
        sf = _build_search_filter(search, ['name', 'ptc_code', 'epic_no', 'mobile', 'assembly', 'district'])
        if sf:
            conditions.append(sf)
    query = _merge_conditions(conditions)

    total = volunteer_requests_col.count_documents(query)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    skip = (page - 1) * per_page

    items = list(volunteer_requests_col.find(query, {'_id': 0}).sort('requested_at', -1).skip(skip).limit(per_page))

    return jsonify({
        'items': items,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@admin_bp.route('/api/volunteer-requests/<ptc_code>/confirm', methods=['POST'])
def confirm_volunteer(ptc_code):
    """Confirm a volunteer request."""
    result = volunteer_requests_col.update_one(
        {'ptc_code': ptc_code, 'status': 'pending'},
        {'$set': {
            'status': 'confirmed',
            'reviewed_at': datetime.now(timezone.utc).isoformat(),
            'reviewed_by': config.ADMIN_USERNAME,
        }}
    )
    if result.modified_count:
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Not found or already reviewed.'}), 404


@admin_bp.route('/api/volunteer-requests/<ptc_code>/reject', methods=['POST'])
def reject_volunteer(ptc_code):
    """Reject a volunteer request."""
    result = volunteer_requests_col.update_one(
        {'ptc_code': ptc_code, 'status': 'pending'},
        {'$set': {
            'status': 'rejected',
            'reviewed_at': datetime.now(timezone.utc).isoformat(),
            'reviewed_by': config.ADMIN_USERNAME,
        }}
    )
    if result.modified_count:
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Not found or already reviewed.'}), 404


@admin_bp.route('/confirmed-volunteers')
def confirmed_volunteers_page():
    """Show confirmed volunteers."""
    return render_template('admin/confirmed_volunteers.html')


@admin_bp.route('/api/confirmed-volunteers')
def api_confirmed_volunteers():
    """JSON API for confirmed volunteers - server-side search & pagination."""
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    conditions = [{'status': 'confirmed'}]
    if search:
        sf = _build_search_filter(search, ['name', 'ptc_code', 'epic_no', 'mobile'])
        if sf:
            conditions.append(sf)
    query = _merge_conditions(conditions)

    total = volunteer_requests_col.count_documents(query)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    skip = (page - 1) * per_page

    items = list(volunteer_requests_col.find(query, {'_id': 0}).sort('reviewed_at', -1).skip(skip).limit(per_page))

    return jsonify({
        'items': items, 'total': total,
        'page': page, 'per_page': per_page, 'total_pages': total_pages,
    })


# ── Booth Agent Requests ─────────────────────────────────────────

@admin_bp.route('/booth-agent-requests')
def booth_agent_requests_page():
    """Show booth agent requests."""
    return render_template('admin/booth_agent_requests.html')


@admin_bp.route('/api/booth-agent-requests')
def api_booth_agent_requests():
    """JSON API for booth agent requests - server-side search & pagination."""
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    conditions = []
    if status:
        conditions.append({'status': status})
    if search:
        sf = _build_search_filter(search, ['name', 'ptc_code', 'epic_no', 'mobile', 'assembly', 'district'])
        if sf:
            conditions.append(sf)
    query = _merge_conditions(conditions)

    total = booth_agent_requests_col.count_documents(query)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    skip = (page - 1) * per_page

    items = list(booth_agent_requests_col.find(query, {'_id': 0}).sort('requested_at', -1).skip(skip).limit(per_page))

    return jsonify({
        'items': items,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@admin_bp.route('/api/booth-agent-requests/<ptc_code>/confirm', methods=['POST'])
def confirm_booth_agent(ptc_code):
    """Confirm a booth agent request."""
    result = booth_agent_requests_col.update_one(
        {'ptc_code': ptc_code, 'status': 'pending'},
        {'$set': {
            'status': 'confirmed',
            'reviewed_at': datetime.now(timezone.utc).isoformat(),
            'reviewed_by': config.ADMIN_USERNAME,
        }}
    )
    if result.modified_count:
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Not found or already reviewed.'}), 404


@admin_bp.route('/api/booth-agent-requests/<ptc_code>/reject', methods=['POST'])
def reject_booth_agent(ptc_code):
    """Reject a booth agent request."""
    result = booth_agent_requests_col.update_one(
        {'ptc_code': ptc_code, 'status': 'pending'},
        {'$set': {
            'status': 'rejected',
            'reviewed_at': datetime.now(timezone.utc).isoformat(),
            'reviewed_by': config.ADMIN_USERNAME,
        }}
    )
    if result.modified_count:
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Not found or already reviewed.'}), 404


@admin_bp.route('/confirmed-booth-agents')
def confirmed_booth_agents_page():
    """Show confirmed booth agents."""
    return render_template('admin/confirmed_booth_agents.html')


@admin_bp.route('/api/confirmed-booth-agents')
def api_confirmed_booth_agents():
    """JSON API for confirmed booth agents - server-side search & pagination."""
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(max(per_page, 5), 100)

    conditions = [{'status': 'confirmed'}]
    if search:
        sf = _build_search_filter(search, ['name', 'ptc_code', 'epic_no', 'mobile'])
        if sf:
            conditions.append(sf)
    query = _merge_conditions(conditions)

    total = booth_agent_requests_col.count_documents(query)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    skip = (page - 1) * per_page

    items = list(booth_agent_requests_col.find(query, {'_id': 0}).sort('reviewed_at', -1).skip(skip).limit(per_page))

    return jsonify({
        'items': items, 'total': total,
        'page': page, 'per_page': per_page, 'total_pages': total_pages,
    })


# Register admin blueprint
app.register_blueprint(admin_bp)

# Register health check blueprint
app.register_blueprint(health_bp)


# ══════════════════════════════════════════════════════════════════
#  RUN
# ══════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Voter ID Card Generator v4.0')
    parser.add_argument('--port', type=int, default=5000)
    parser.add_argument('--host', type=str, default='127.0.0.1')
    parser.add_argument('--debug', action='store_true')
    args = parser.parse_args()

    print("=" * 60)
    print("  VOTER ID CARD GENERATOR v4.0")
    print(f"  User  : http://{args.host}:{args.port}/")
    print(f"  Admin : http://{args.host}:{args.port}/admin")
    print("  Database: MongoDB Atlas | Photos: Cloudinary")
    print("=" * 60)

    app.run(host=args.host, port=args.port, debug=args.debug)
